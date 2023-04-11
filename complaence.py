import openpyxl
from datetime import datetime


class TimeOver:
    def __init__(self, path):
        self.date_now = datetime.now()
        self.workers = {}
        self.path = path


    def download_from_excel(self):
        wb = openpyxl.load_workbook(f'{self.path}\\учет данных.xlsx')
        worksheet = wb['учет']
        for col in worksheet.iter_cols(1, worksheet.max_column):
            for row in range(0, 1):
                self.workers[col[row].value] = []
            for rows in range(1, worksheet.max_row):
                self.workers[col[row].value].append(col[rows].value)
        del wb['учет']
        wb.create_sheet('учет')
        wb.save(f'{self.path}\\учет данных.xlsx')
        wb.close()
        return self.workers
    

    def calculating_the_length_of_service(self):
        for idx in range(len(self.workers['Имя'])):
            if self.workers['Дата начала работы по ИБ(без )'][idx] != None:
                if type(self.workers['Дата начала работы по ИБ(без )'][idx]) == datetime:
                    date_work_before_cft = self.workers['Дата начала работы по ИБ(без )'][idx]
                else:
                    date_work_before_cft = (datetime.strptime(self.workers['Дата начала работы по ИБ(без )'][idx], '%d.%m.%Y')).date()
            else:
                date_work_before_cft = None
            if type(self.workers['Дата начала работы по ИБ(в )'][idx]) == datetime:
                date_work_in_the_cft = self.workers['Дата начала работы по ИБ(в )'][idx]
            else:
                date_work_in_the_cft = (datetime.strptime(self.workers['Дата начала работы по ИБ(в )'][idx], '%d.%m.%Y')).date()
            if date_work_before_cft == None:
                delta_date_work_in_the_cft_and_now = self.date_now - date_work_in_the_cft
                years = ((delta_date_work_in_the_cft_and_now.days) // 365)
                months = ((delta_date_work_in_the_cft_and_now.days) % 365) // 30
                days = ((delta_date_work_in_the_cft_and_now.days) % 365) % 30
            else:
                delta_date_work_in_the_cft_and_now = self.date_now - date_work_in_the_cft
                delta_date_work_before_cft_and_date_work_in_the_cft = date_work_in_the_cft - date_work_before_cft
                years = (delta_date_work_in_the_cft_and_now.days + delta_date_work_before_cft_and_date_work_in_the_cft.days) // 365
                months = ((delta_date_work_in_the_cft_and_now.days + delta_date_work_before_cft_and_date_work_in_the_cft.days) % 365) // 30
                days = ((delta_date_work_in_the_cft_and_now.days + delta_date_work_before_cft_and_date_work_in_the_cft.days) % 365) % 30
            
            if 1 < (years) % 10 <= 4:
                if 1 < months <= 4:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} года, {months} месяца, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяца, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяца, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяца, {days} дней'
                elif months == 1:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} года, {months} месяц, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяц, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяц, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяц, {days} дней'
                else:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} года, {months} месяцев, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяцев, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяцев, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} года, {months} месяцев, {days} дней'
            
            elif (years) % 10 < 1:
                if 1 < months <= 4:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{months} месяца, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{months} месяца, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{months} месяца, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{months} месяца,{days} дней'
                elif months == 1:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{months} месяц, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{months} месяц, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{months} месяц, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{months} месяц, {days} дней'
                else:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{months} месяцев, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{months} месяцев, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{months} месяцев, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{months} месяцев, {days} дней'
            
            elif (years) % 10 == 1:
                if 1 < months <= 4:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} год, {months} месяца, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяца, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяца, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяца, {days} дней'
                elif months == 1:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} год, {months} месяц, {days} дней'
                    else:   
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяц, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяц, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяц, {days} дней'
                else:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} год, {months} месяцев, {days} дней'
                    else:                   
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяцев, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяцев, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} год, {months} месяцев, {days} дней'
            
            else:
                if 1 < months <= 4:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяца, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяца, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяца, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяца, {days} дней'
                elif months == 1:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяц, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяц, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяц, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяц, {days} дней'
                else:
                    if 10 <= days % 100 <= 20:
                        self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяцев, {days} дней'
                    else:
                        if 1 < (days) % 10 <= 4:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяцев, {days} дня'
                        elif (days) % 10 == 1:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяцев, {days} день'
                        else:
                            self.workers['Опыт работы'][idx] = f'{years} лет, {months} месяцев, {days} дней'


    def writer_on_the_sheet(self, sheet, params):
        sheet['A1'] = 'Имя'
        sheet['B1'] = 'Фамилия'
        sheet['C1'] = 'Должность'
        sheet['D1'] = 'Отдел'
        sheet['E1'] = 'Дата начала работы по ИБ(без )'
        sheet['F1'] = 'Дата начала работы по ИБ(в )'
        sheet['G1'] = 'Опыт работы'
        sheet['H1'] = 'Увольнение'
        sheet.freeze_panes = 'A2'  
        sheet.append(params)
        return None

        
    def add_on_sheet(self):
        wb = openpyxl.load_workbook(f'{self.path}\\учет данных.xlsx')
        sheet = wb['учет']
        for idx in range(len(self.workers['Имя'])):
            if type(self.workers['Дата начала работы по ИБ(в )'][idx]) == datetime:
                date_work_in_the_cft = self.workers['Дата начала работы по ИБ(в )'][idx]
            else:
                date_work_in_the_cft = (self.workers['Дата начала работы по ИБ(в )'][idx]).strftime('%d.%m.%Y')
            if self.workers['Дата начала работы по ИБ(без )'][idx] != None:
                if type(self.workers['Дата начала работы по ИБ(без )'][idx]) == datetime:
                    date_work_before_cft = self.workers['Дата начала работы по ИБ(без )'][idx]
                else:
                    date_work_before_cft = (self.workers['Дата начала работы по ИБ(без )'][idx]).strftime('%d.%m.%Y')
            else:
                date_work_before_cft = None
            params = [
                self.workers['Имя'][idx],
                self.workers['Фамилия'][idx],
                self.workers['Должность'][idx],
                self.workers['Отдел'][idx],
                date_work_before_cft,
                date_work_in_the_cft,
                self.workers['Опыт работы'][idx],
                self.workers['Увольнение'][idx]
            ]
            self.writer_on_the_sheet(sheet, params)
        wb.save(f'{self.path}\\учет данных.xlsx')
        wb.close()
        return None


if __name__ == '__main__':
    model = TimeOver('J:\code\complaence_35')
    model.download_from_excel()
    model.calculating_the_length_of_service()
    model.add_on_sheet()
